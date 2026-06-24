"""
Export utilities for PDF, Excel, and CSV generation.
"""

import csv
import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_to_pdf(title, headers, data, filename=None):
    """Generate a PDF report.

    Args:
        title: Report title
        headers: List of column headers
        data: List of rows (each row is a list of values)
        filename: Optional filename

    Returns:
        BytesIO buffer containing the PDF
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#0F172A'),
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#6B7280'),
    )

    elements = []

    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 12))

    # Table
    table_data = [headers] + data
    table = Table(table_data, repeatRows=1)

    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor('#F8FAFC'),
        ]),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_to_excel(title, headers, data, filename=None):
    """Generate an Excel report.

    Args:
        title: Report title (used as sheet name)
        headers: List of column headers
        data: List of rows

    Returns:
        BytesIO buffer containing the Excel file
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='0F172A')
    title_cell.alignment = Alignment(horizontal='center')

    # Date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(
        row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
    )
    date_cell.font = Font(name='Calibri', size=10, color='6B7280')
    date_cell.alignment = Alignment(horizontal='center')

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for row_idx, row in enumerate(data, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if (row_idx - 5) % 2 == 1:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(4, len(data) + 5)
        )
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = min(max_length + 4, 40)

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_to_csv(headers, data):
    """Generate a CSV report.

    Args:
        headers: List of column headers
        data: List of rows

    Returns:
        BytesIO buffer containing the CSV
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(data)

    byte_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
    byte_buffer.seek(0)
    return byte_buffer
